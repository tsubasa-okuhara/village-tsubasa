"""
経費精算・レシート管理 Streamlit アプリケーション
電子帳簿保存法対応 / マルチユーザー対応

使用方法: streamlit run app.py
"""

import os
import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from database import (
    init_db, insert_receipt, update_receipt, soft_delete_receipt,
    search_receipts, get_receipt_by_id, get_audit_log,
    get_categories, get_unique_vendors, get_receipt_stats,
    sign_in_helper, sign_up_helper, sign_out_helper,
    is_admin, search_all_receipts, get_all_receipt_stats,
)
from image_utils import validate_image, compute_hash, save_image, load_image_for_display
from ocr_utils import extract_receipt_from_image

# ===== ページ設定 =====
# スマホ対応: サイドバーを初期非表示、縦積みレイアウトのため layout=centered
st.set_page_config(
    page_title="経費精算・レシート管理",
    page_icon="🧾",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# ===== スタイル設定 =====
st.markdown("""
<style>
    .metric-box {
        background-color: #f0f2f6;
        padding: 15px;
        border-radius: 8px;
        margin: 10px 0;
    }
    .success-box {
        background-color: #d4edda;
        padding: 10px;
        border-radius: 5px;
        color: #155724;
    }
    .error-box {
        background-color: #f8d7da;
        padding: 10px;
        border-radius: 5px;
        color: #721c24;
    }
</style>
""", unsafe_allow_html=True)

# ===== 初期化処理 =====
if 'initialized' not in st.session_state:
    init_db()
    st.session_state.initialized = True


# ===== 認証 (Supabase Auth + helper_master ホワイトリスト) =====

def _get_current_user() -> dict:
    """現在のログインユーザー情報 (未ログインなら None)"""
    return st.session_state.get("current_user")


def login_page() -> bool:
    """
    ログイン / 新規登録ページ。
    成功時は st.session_state['current_user'] に helper 情報を保存して True を返す。
    未ログインの間は False を返す。
    """
    if _get_current_user():
        return True

    st.title("🔐 経費精算・レシート管理")
    st.caption("電子帳簿保存法対応システム / ヘルパー専用")
    st.write("")

    tab_login, tab_signup = st.tabs(["🔑 ログイン", "📝 新規登録"])

    # --- ログインタブ ---
    with tab_login:
        st.write("登録済みのヘルパーさんはこちらからログインしてください。")
        with st.form("login_form"):
            login_email = st.text_input("メールアドレス", key="login_email")
            login_password = st.text_input("パスワード", type="password", key="login_password")
            submit_login = st.form_submit_button("ログイン", use_container_width=True, type="primary")

        if submit_login:
            result = sign_in_helper(login_email, login_password)
            if result["ok"]:
                helper = result["helper"] or {}
                st.session_state["current_user"] = {
                    "helper_email": (login_email or "").strip().lower(),
                    "helper_name": helper.get("helper_name", ""),
                }
                st.rerun()
            else:
                st.error(f"❌ {result['error']}")

    # --- 新規登録タブ ---
    with tab_signup:
        st.write("初めて利用する方は、こちらからアカウントを作成してください。")
        st.caption("※ 管理者がヘルパーマスタに登録済みのメールアドレスのみ登録できます。")
        with st.form("signup_form"):
            signup_email = st.text_input("メールアドレス", key="signup_email")
            signup_password = st.text_input(
                "パスワード (8文字以上)",
                type="password",
                key="signup_password",
            )
            signup_password_confirm = st.text_input(
                "パスワード（確認）",
                type="password",
                key="signup_password_confirm",
            )
            submit_signup = st.form_submit_button("アカウント作成", use_container_width=True, type="primary")

        if submit_signup:
            if signup_password != signup_password_confirm:
                st.error("❌ パスワードが一致しません")
            else:
                result = sign_up_helper(signup_email, signup_password)
                if result["ok"]:
                    st.success("✅ アカウントを作成しました。ログインタブからログインしてください。")
                else:
                    st.error(f"❌ {result['error']}")

    return False


def format_currency(amount: float) -> str:
    """金額を通貨形式でフォーマット"""
    return f"¥{amount:,.0f}"


def create_excel_download(
    results: list,
    date_from=None,
    date_to=None,
    include_helper: bool = False,
) -> bytes:
    """
    検索結果をExcelファイルに変換して返す
    include_helper=True にすると「登録者」列を追加 (管理者用・税理士提出用)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "経費明細"

    # スタイル定義
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    title_font = Font(name="Arial", bold=True, size=14)
    subtitle_font = Font(name="Arial", size=10, color="666666")
    currency_format = '#,##0'
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # ヘッダー定義 (費目列を追加)
    if include_helper:
        headers = ["No.", "取引日", "費目", "金額", "取引先", "登録者"]
        col_widths = [6, 14, 14, 14, 30, 20]
    else:
        headers = ["No.", "取引日", "費目", "金額", "取引先"]
        col_widths = [6, 14, 14, 14, 30]
    last_col_letter = get_column_letter(len(headers))

    # タイトル行
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = "経費精算明細書"
    ws["A1"].font = title_font

    # 期間表示
    period_text = ""
    if date_from and date_to:
        period_text = f"期間: {date_from} ～ {date_to}"
    elif date_from:
        period_text = f"期間: {date_from} ～"
    elif date_to:
        period_text = f"期間: ～ {date_to}"
    if period_text:
        ws.merge_cells(f"A2:{last_col_letter}2")
        ws["A2"] = period_text
        ws["A2"].font = subtitle_font

    # 出力日
    ws["A3"] = f"出力日: {datetime.now().strftime('%Y年%m月%d日')}"
    ws["A3"].font = subtitle_font

    # ヘッダー行（5行目から）
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 金額列のインデックス (1始まり)
    amount_col_idx = 4

    # データ行
    total_amount = 0
    for row_idx, receipt in enumerate(results, 1):
        row = row_idx + 5
        # No.
        no_cell = ws.cell(row=row, column=1, value=row_idx)
        no_cell.border = thin_border
        no_cell.alignment = Alignment(horizontal="center")
        # 取引日
        ws.cell(row=row, column=2, value=receipt['transaction_date']).border = thin_border
        # 費目
        ws.cell(row=row, column=3, value=receipt.get('category') or '').border = thin_border
        # 金額
        amount_cell = ws.cell(row=row, column=amount_col_idx, value=receipt['amount'])
        amount_cell.number_format = currency_format
        amount_cell.border = thin_border
        amount_cell.alignment = Alignment(horizontal="right")
        # 取引先
        ws.cell(row=row, column=5, value=receipt.get('vendor') or '').border = thin_border
        # 登録者 (管理者出力のみ)
        if include_helper:
            helper_display = receipt.get('helper_name') or receipt.get('helper_email') or ''
            ws.cell(row=row, column=6, value=helper_display).border = thin_border
        total_amount += receipt['amount']

    # 合計行
    total_row = len(results) + 6
    label_cell = ws.cell(row=total_row, column=amount_col_idx - 1, value="合計")
    label_cell.font = Font(bold=True)
    label_cell.alignment = Alignment(horizontal="right")
    label_cell.border = thin_border
    total_cell = ws.cell(row=total_row, column=amount_col_idx, value=total_amount)
    total_cell.font = Font(bold=True)
    total_cell.number_format = currency_format
    total_cell.alignment = Alignment(horizontal="right")
    total_cell.border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="double", color="4472C4"),
        bottom=Side(style="double", color="4472C4"),
    )

    # バッファに書き出し
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ===== ページ1: レシート登録 =====
def page_receipt_registration(user: dict):
    st.title("📝 レシート登録")

    helper_email = user["helper_email"]
    helper_name = user.get("helper_name", "")

    st.subheader("ステップ1: 画像アップロード")
    uploaded_file = st.file_uploader(
        "レシート画像をアップロード（JPG, PNG）",
        type=["jpg", "jpeg", "png"],
        help="スマホの場合: タップするとカメラ or 写真から選べます"
    )

    # ===== 画像バリデーション =====
    image_data = None
    ocr_result = None

    if uploaded_file:
        # ファイルを読み込む
        image_bytes = uploaded_file.getvalue()
        validation = validate_image(uploaded_file)

        # 画像プレビュー (1カラム)
        st.image(image_bytes, caption="アップロード画像", width="stretch")

        # 検証結果 (折りたたみで詳細、バッジで概要のみ)
        if validation['is_valid']:
            st.success("✅ 画像検証: 合格")
        else:
            st.error("❌ 画像検証: 不合格")

        with st.expander("画像の詳細を見る", expanded=False):
            st.write(f"- 形式: {validation['format']}")
            st.write(f"- サイズ: {validation['width']}×{validation['height']}px")
            st.write(f"- 色モード: {validation['color_mode']}")
            st.write(f"- DPI: {validation['dpi']}")

            # DPIチェック
            if validation['dpi'] and validation['dpi'] >= 200:
                st.success(f"✅ DPI要件OK（{validation['dpi']}dpi ≥ 200dpi）")
            else:
                st.error(f"❌ DPI不足（{validation['dpi']}dpi < 200dpi）")

            # 色モードチェック
            if validation['color_mode'] in ['RGB', 'RGBA']:
                st.success(f"✅ 色モードOK（{validation['color_mode']}）")
            else:
                st.error(f"❌ 色モード不正（{validation['color_mode']}）")

        # エラー・警告表示
        if validation['errors']:
            st.error("**エラー:**")
            for error in validation['errors']:
                st.write(f"- {error}")

        if validation['warnings']:
            st.warning("**警告:**")
            for warning in validation['warnings']:
                st.write(f"- {warning}")

        # 検証成功時の処理
        if validation['is_valid']:
            image_data = {
                'bytes': image_bytes,  # 保存は元画像（法的要件）
                'hash': compute_hash(image_bytes),
                'dpi': validation['dpi'],
                'color_mode': validation['color_mode']
            }

            # ===== Claude Vision による情報抽出 =====
            st.subheader("ステップ2: AI情報抽出 (Claude Vision)")

            with st.spinner("Claude が画像を読み取っています..."):
                ocr_result = extract_receipt_from_image(image_bytes)

            confidence = ocr_result.get('confidence_score', 0)
            note = ocr_result.get('confidence_note', '') or ''

            # 信頼度に応じて色を変える
            if confidence >= 80:
                st.success(f"✅ 抽出成功 (信頼度: {confidence}/100)")
            elif confidence >= 50:
                st.warning(f"⚠️ 一部不確実 (信頼度: {confidence}/100) — 下のフォームで必ず確認してください")
            else:
                st.error(f"❌ 抽出精度が低い (信頼度: {confidence}/100) — 手入力で確認・修正してください")

            if note:
                st.caption(f"💬 {note}")

            st.write("**抽出結果:**")
            info_lines = []
            info_lines.append(f"- 店舗名: {ocr_result.get('store_name') or '検出不可'}")
            info_lines.append(f"- 取引日: {ocr_result.get('transaction_date') or '検出不可'}")
            info_lines.append(
                f"- 合計金額: {format_currency(ocr_result['total_amount']) if ocr_result.get('total_amount') else '検出不可'}"
            )
            if ocr_result.get('phone_number'):
                info_lines.append(f"- 電話番号: {ocr_result['phone_number']}")
            if ocr_result.get('invoice_number'):
                info_lines.append(f"- インボイス番号: {ocr_result['invoice_number']}")
            if ocr_result.get('address'):
                info_lines.append(f"- 住所: {ocr_result['address']}")
            st.info("\n".join(info_lines))

            if st.checkbox("抽出結果の詳細 (JSON) を表示"):
                st.code(ocr_result.get('raw_text', ''), language='json')

    # ===== レシート登録フォーム =====
    if image_data and ocr_result:
        st.subheader("ステップ3: レシート情報入力")

        with st.form("receipt_form"):
            # スマホ対応: 1カラム縦積み
            transaction_date = st.date_input(
                "取引年月日",
                value=datetime.strptime(ocr_result['date'], '%Y-%m-%d').date() if ocr_result['date'] else datetime.now().date(),
                format="YYYY-MM-DD"
            )

            amount = st.number_input(
                "金額",
                value=int(ocr_result['amount']) if ocr_result['amount'] else 0,
                min_value=0,
                step=1,
                format="%d"
            )

            vendors = get_unique_vendors(helper_email)
            vendor_default = ocr_result['vendor'] if ocr_result['vendor'] and ocr_result['vendor'] in vendors else (vendors[0] if vendors else "")

            vendor = st.selectbox(
                "取引先",
                options=vendors + [""] if vendors else [""],
                index=vendors.index(vendor_default) if vendor_default and vendor_default in vendors else len(vendors),
                key="vendor_select"
            )

            # 新規取引先入力
            if vendor == "":
                vendor = st.text_input("取引先を入力", value=ocr_result['vendor'] or "")

            category = st.selectbox(
                "費目",
                options=get_categories(),
                index=0
            )

            # AI が抽出した追加情報 (インボイス番号・電話番号) を初期値にして編集可能に
            phone_number = st.text_input(
                "電話番号 (任意)",
                value=ocr_result.get('phone_number') or "",
                help="AIが抽出した店舗電話番号。必要に応じて修正してください。"
            )

            invoice_number = st.text_input(
                "インボイス登録番号 (任意)",
                value=ocr_result.get('invoice_number') or "",
                help="適格請求書発行事業者登録番号 (T + 13桁)。電子帳簿保存法対応のため記録されます。"
            )

            store_address = st.text_input(
                "店舗住所 (任意)",
                value=ocr_result.get('address') or "",
                help="AIが抽出した店舗住所。必要に応じて修正してください。"
            )

            description = st.text_area(
                "備考",
                value="",
                height=100
            )

            # 登録ボタン
            submit = st.form_submit_button("レシート登録", use_container_width=True, type="primary")

            if submit:
                # バリデーション
                if not transaction_date:
                    st.error("取引年月日を入力してください")
                elif amount <= 0:
                    st.error("金額を入力してください")
                elif not vendor:
                    st.error("取引先を入力してください")
                elif not category:
                    st.error("費目を選択してください")
                else:
                    try:
                        # レシート登録
                        receipt_id = insert_receipt(
                            transaction_date=transaction_date.isoformat(),
                            amount=float(amount),
                            vendor=vendor,
                            category=category,
                            description=description,
                            image_path=save_image(image_data['bytes'], 0),  # 仮ID
                            image_hash=image_data['hash'],
                            image_dpi=image_data['dpi'],
                            image_color_mode=image_data['color_mode'],
                            helper_email=helper_email,
                            helper_name=helper_name,
                            ocr_raw_text=ocr_result.get('raw_text'),
                            created_by=helper_email,
                            phone_number=phone_number or None,
                            invoice_number=invoice_number or None,
                            store_address=store_address or None,
                            ai_confidence=ocr_result.get('confidence_score'),
                        )

                        st.success(f"✅ レシート登録完了！（ID: {receipt_id}）")
                        st.balloons()

                    except Exception as e:
                        st.error(f"エラー: {str(e)}")


# ===== ページ1.5: 手入力（テンキー風 + エクセル風 2タブ） =====
def page_manual_entry(user: dict):
    """
    レシート画像を使わず、手入力でレシートを登録するページ。
    上部 2 タブ:
      - ⌨️ テンキー風: マネーフォワード風の電卓 UI で 1 件ずつ入力
      - 📊 エクセル風: 表形式で複数件をまとめて入力 (st.data_editor)
    画像が無い前提で insert_receipt を呼ぶ。電子帳簿保存法のスキャナ保存対応にはならないため、
    画面上に「紙レシートは別途保管してください」の注意書きを表示する。
    """
    st.title("⌨️ 手入力でレシート登録")
    st.caption(
        "レシート画像を使わずに、手入力で経費を登録できます。"
        "好みの方法を選んでください（途中でタブを切り替えても入力中の内容は保持されません）。"
    )
    st.info(
        "💡 **電子帳簿保存法について**：手入力モードはスキャナ保存要件には対応しません。"
        "紙のレシートは別途保管していただき、AIで読み取りたい場合は「📝 レシート登録」ページから画像をアップロードしてください。"
    )

    helper_email = user["helper_email"]
    helper_name = user.get("helper_name", "")

    # カテゴリーをページ内で1度だけ取得（Supabase 呼び出し節約）
    categories = get_categories()
    if not categories:
        st.warning("⚠️ 費目マスタが空です。管理者に連絡してください。")
        return

    tab_keypad, tab_excel = st.tabs(["⌨️ テンキー風（1件ずつ）", "📊 エクセル風（複数行まとめて）"])

    # ============================================================
    # タブ A: テンキー風入力
    # ============================================================
    with tab_keypad:
        # セッション状態の初期化
        if "manual_amount_str" not in st.session_state:
            st.session_state.manual_amount_str = ""

        amount_int = int(st.session_state.manual_amount_str or "0")

        # 大きな金額表示
        st.markdown(
            f"""
            <div style="background:#f7f9fc; border-radius:12px; padding:18px 12px;
                        margin:8px 0 12px; text-align:center;
                        border:1px solid #e4e8ee;">
                <div style="color:#6b7785; font-size:0.85rem; letter-spacing:1px;">金額</div>
                <div style="color:#1a2733; font-size:2.6rem; font-weight:700; line-height:1.1; margin-top:4px;">
                    ¥{amount_int:,}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        # テンキー（3列 × 4行）
        keypad_rows = [
            ["7", "8", "9"],
            ["4", "5", "6"],
            ["1", "2", "3"],
            ["C", "0", "⌫"],
        ]
        for row in keypad_rows:
            cols = st.columns(3)
            for i, key in enumerate(row):
                clicked = cols[i].button(
                    key,
                    use_container_width=True,
                    key=f"keypad_btn_{key}",
                )
                if clicked:
                    if key == "C":
                        st.session_state.manual_amount_str = ""
                    elif key == "⌫":
                        st.session_state.manual_amount_str = st.session_state.manual_amount_str[:-1]
                    else:
                        # 先頭ゼロ抑止 + 桁数上限 9 (¥999,999,999)
                        if st.session_state.manual_amount_str == "" and key == "0":
                            pass
                        elif len(st.session_state.manual_amount_str) < 9:
                            st.session_state.manual_amount_str += key
                    st.rerun()

        st.divider()

        # その他のフィールド（フォーム）
        with st.form("keypad_entry_form", clear_on_submit=False):
            kp_date = st.date_input(
                "取引年月日",
                value=datetime.now().date(),
                format="YYYY-MM-DD",
                key="keypad_date",
            )

            kp_category = st.selectbox(
                "費目",
                options=categories,
                index=0,
                key="keypad_category",
            )

            existing_vendors = get_unique_vendors(helper_email)
            kp_vendor_select = st.selectbox(
                "取引先（既存から選ぶ）",
                options=[""] + existing_vendors,
                index=0,
                key="keypad_vendor_select",
                help="新しい取引先は下のテキスト欄に入力してください",
            )
            kp_vendor_text = st.text_input(
                "取引先（新規入力 or 修正）",
                value="",
                key="keypad_vendor_text",
                placeholder="例: スターバックス渋谷店",
            )

            kp_description = st.text_area(
                "備考（任意）",
                value="",
                height=80,
                key="keypad_description",
            )

            kp_submit = st.form_submit_button(
                "💾 この内容で登録",
                use_container_width=True,
                type="primary",
            )

            if kp_submit:
                amount_now = int(st.session_state.manual_amount_str or "0")
                vendor_final = (kp_vendor_text.strip() or kp_vendor_select.strip())

                if amount_now <= 0:
                    st.error("金額を入力してください（テンキーで0より大きい数値）")
                elif not vendor_final:
                    st.error("取引先を入力してください")
                else:
                    try:
                        rid = insert_receipt(
                            transaction_date=kp_date.isoformat(),
                            amount=float(amount_now),
                            vendor=vendor_final,
                            category=kp_category,
                            description=kp_description,
                            helper_email=helper_email,
                            helper_name=helper_name,
                            created_by=helper_email,
                        )
                        st.success(f"✅ 登録完了！（ID: {rid}）")
                        st.balloons()
                        # リセット
                        st.session_state.manual_amount_str = ""
                        st.rerun()
                    except Exception as e:
                        st.error(f"エラー: {str(e)}")

    # ============================================================
    # タブ B: エクセル風入力（複数件まとめて）
    # ============================================================
    with tab_excel:
        st.write(
            "下の表に直接打ち込んで、最後に **「まとめて登録」** ボタンを押してください。"
            "**金額が0** または **費目が空** の行は自動的にスキップされます。"
        )
        st.caption("💡 行を追加・削除するには表の右端（… メニュー）または末尾の空行を活用してください。")

        # st.data_editor の key を変えることで「登録後にリセット」が実現できる
        if "excel_reset_count" not in st.session_state:
            st.session_state.excel_reset_count = 0

        default_rows = 5
        today = datetime.now().date()
        initial_df = pd.DataFrame({
            "取引日": [today] * default_rows,
            "金額": [0] * default_rows,
            "費目": [categories[0]] * default_rows if categories else [""] * default_rows,
            "取引先": [""] * default_rows,
            "備考": [""] * default_rows,
        })

        edited_df = st.data_editor(
            initial_df,
            num_rows="dynamic",
            use_container_width=True,
            key=f"excel_editor_{st.session_state.excel_reset_count}",
            column_config={
                "取引日": st.column_config.DateColumn(
                    "取引日",
                    format="YYYY-MM-DD",
                    required=True,
                ),
                "金額": st.column_config.NumberColumn(
                    "金額（円）",
                    min_value=0,
                    step=1,
                    format="%d",
                ),
                "費目": st.column_config.SelectboxColumn(
                    "費目",
                    options=categories,
                    required=True,
                ),
                "取引先": st.column_config.TextColumn(
                    "取引先",
                    max_chars=100,
                ),
                "備考": st.column_config.TextColumn(
                    "備考",
                    max_chars=200,
                ),
            },
        )

        # 集計プレビュー
        try:
            valid_preview = edited_df[
                (edited_df["金額"].fillna(0).astype(float) > 0)
                & edited_df["費目"].fillna("").astype(str).str.len().gt(0)
            ]
            preview_total = float(valid_preview["金額"].fillna(0).sum())
            preview_count = len(valid_preview)
            c1, c2 = st.columns(2)
            c1.metric("登録される件数", f"{preview_count}件")
            c2.metric("合計金額", format_currency(preview_total))
        except Exception:
            pass

        if st.button(
            "📥 上記の内容をまとめて登録",
            use_container_width=True,
            type="primary",
        ):
            success_count = 0
            skipped_count = 0
            errors = []

            for idx, row in edited_df.iterrows():
                try:
                    amt = float(row["金額"]) if pd.notna(row["金額"]) else 0.0
                except (ValueError, TypeError):
                    amt = 0.0
                cat = str(row["費目"]).strip() if pd.notna(row["費目"]) else ""
                date_val = row["取引日"]

                # スキップ条件: 金額0以下 / 費目空 / 日付欠損
                if amt <= 0 or not cat or pd.isna(date_val):
                    skipped_count += 1
                    continue

                try:
                    # date_val が datetime.date / pandas.Timestamp / str の可能性
                    if hasattr(date_val, "isoformat"):
                        date_str = date_val.isoformat()[:10]
                    else:
                        date_str = str(date_val)[:10]

                    rid = insert_receipt(
                        transaction_date=date_str,
                        amount=amt,
                        vendor=str(row["取引先"]).strip() if pd.notna(row["取引先"]) else "",
                        category=cat,
                        description=str(row["備考"]).strip() if pd.notna(row["備考"]) else "",
                        helper_email=helper_email,
                        helper_name=helper_name,
                        created_by=helper_email,
                    )
                    success_count += 1
                except Exception as e:
                    errors.append(f"{idx + 1}行目: {str(e)}")

            if success_count:
                st.success(f"✅ {success_count}件を登録しました")
                st.balloons()
                # 表をリセット
                st.session_state.excel_reset_count += 1
                st.rerun()
            if skipped_count:
                st.info(f"ℹ️ {skipped_count}件は金額0または費目空のためスキップしました")
            if errors:
                st.error("以下の行で登録エラー:\n\n" + "\n\n".join(errors))
            if not success_count and not skipped_count and not errors:
                st.warning("登録対象の行がありませんでした。表に金額と費目を入力してください。")


# ===== ページ2: 検索・一覧 =====
def page_search_and_list(user: dict):
    st.title("🔍 検索・一覧")

    helper_email = user["helper_email"]

    # 検索パネル (スマホ対応: 1カラム縦積み)
    with st.expander("🔎 検索条件", expanded=True):
        date_from = st.date_input(
            "取引年月日：開始",
            value=datetime.now().date() - timedelta(days=90),
            format="YYYY-MM-DD"
        )

        date_to = st.date_input(
            "取引年月日：終了",
            value=datetime.now().date(),
            format="YYYY-MM-DD"
        )

        vendors = [""] + get_unique_vendors(helper_email)
        vendor = st.selectbox(
            "取引先",
            options=vendors,
            index=0
        )

        amount_min = st.number_input(
            "金額：最小",
            value=0,
            min_value=0,
            step=100,
            format="%d"
        )

        amount_max = st.number_input(
            "金額：最大",
            value=999999,
            min_value=0,
            step=100,
            format="%d"
        )

    # 検索実行
    if st.button("検索", use_container_width=True, type="primary"):
        results = search_receipts(
            helper_email=helper_email,
            date_from=date_from.isoformat() if date_from else None,
            date_to=date_to.isoformat() if date_to else None,
            amount_min=amount_min if amount_min > 0 else None,
            amount_max=amount_max if amount_max < 999999 else None,
            vendor=vendor if vendor else None
        )

        if results:
            # 統計情報 (スマホでも3列の方がコンパクトなので残す)
            stats = get_receipt_stats(
                helper_email=helper_email,
                date_from=date_from.isoformat(),
                date_to=date_to.isoformat()
            )

            c1, c2, c3 = st.columns(3)
            c1.metric("件数", f"{stats['count']}件")
            c2.metric("合計", format_currency(stats['total']))
            c3.metric("平均", format_currency(stats['total'] / stats['count'] if stats['count'] > 0 else 0))

            # 結果テーブル
            st.subheader("検索結果")

            # DataFrameに変換
            df = pd.DataFrame(results)
            df_display = df[[
                'id', 'transaction_date', 'amount', 'vendor', 'category', 'created_at'
            ]].copy()

            df_display.columns = ['ID', '取引日', '金額', '取引先', '費目', '登録日']
            df_display['金額'] = df_display['金額'].apply(lambda x: format_currency(x))

            st.dataframe(
                df_display,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "ID": st.column_config.NumberColumn(width=50),
                    "取引日": st.column_config.TextColumn(width=100),
                    "金額": st.column_config.TextColumn(width=100),
                    "取引先": st.column_config.TextColumn(width=150),
                    "費目": st.column_config.TextColumn(width=100),
                    "登録日": st.column_config.TextColumn(width=150),
                }
            )

            # Excel出力ボタン
            st.subheader("📥 Excel出力")
            excel_data = create_excel_download(
                results,
                date_from=date_from.isoformat() if date_from else None,
                date_to=date_to.isoformat() if date_to else None,
            )
            filename = f"経費明細_{datetime.now().strftime('%Y%m%d')}.xlsx"
            st.download_button(
                label="Excelファイルをダウンロード",
                data=excel_data,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

            # 詳細表示用
            st.subheader("詳細確認")
            selected_id = st.selectbox(
                "レシートを選択",
                options=[r['id'] for r in results],
                format_func=lambda x: f"ID: {x}"
            )

            if selected_id:
                st.session_state.selected_receipt_id = selected_id
                st.info("「詳細・編集」ページで詳細を表示・編集できます")

        else:
            st.info("該当するレシートがありません")


# ===== ページ3: 詳細・編集 =====
def page_detail_and_edit(user: dict):
    st.title("📋 詳細・編集")

    helper_email = user["helper_email"]

    # レシート選択 (スマホ対応: 縦積み)
    all_receipts = search_receipts(helper_email=helper_email)
    if all_receipts:
        receipt_id = st.selectbox(
            "レシートを選択",
            options=[r['id'] for r in all_receipts],
            format_func=lambda x: f"ID: {x} - {next((r['transaction_date'] for r in all_receipts if r['id'] == x), 'N/A')}"
        )
    else:
        st.info("登録されたレシートがありません")
        return

    if st.button("🔄 再読み込み", use_container_width=True):
        st.rerun()

    # レシート詳細を取得
    receipt = get_receipt_by_id(receipt_id, helper_email=helper_email)

    if not receipt:
        st.error("レシートが見つかりません")
        return

    # 画像表示
    st.subheader("レシート画像")

    if receipt['image_path']:
        image = load_image_for_display(receipt['image_path'])
        if image:
            st.image(image, caption="スキャン画像", width="stretch")
        else:
            st.warning("画像が見つかりません")

    # 現在の情報を表示 (スマホ対応: 縦積み)
    st.subheader("登録情報")
    st.write(f"**ID:** {receipt['id']}")
    st.write(f"**取引日:** {receipt['transaction_date']}")
    st.write(f"**金額:** {format_currency(receipt['amount'])}")
    st.write(f"**取引先:** {receipt['vendor']}")
    st.write(f"**費目:** {receipt['category']}")
    st.write(f"**登録日:** {receipt['created_at']}")

    # AI抽出情報 (存在する場合のみ表示)
    if receipt.get('phone_number'):
        st.write(f"**電話番号:** {receipt['phone_number']}")
    if receipt.get('invoice_number'):
        st.write(f"**インボイス番号:** {receipt['invoice_number']}")
    if receipt.get('store_address'):
        st.write(f"**店舗住所:** {receipt['store_address']}")
    if receipt.get('ai_confidence') is not None:
        st.write(f"**AI信頼度:** {receipt['ai_confidence']}/100")

    if receipt['description']:
        st.write(f"**備考:** {receipt['description']}")

    # OCRテキスト表示
    if receipt['ocr_raw_text'] and st.checkbox("OCRテキストを表示"):
        st.text_area("抽出テキスト", value=receipt['ocr_raw_text'], height=150, disabled=True)

    # ===== 編集フォーム =====
    st.divider()
    st.subheader("情報を編集")

    with st.form("edit_form"):
        # スマホ対応: 1カラム縦積み
        new_date = st.date_input(
            "取引年月日",
            value=datetime.strptime(receipt['transaction_date'], '%Y-%m-%d').date(),
            format="YYYY-MM-DD"
        )

        new_amount = st.number_input(
            "金額",
            value=int(receipt['amount']),
            min_value=0,
            step=1,
            format="%d"
        )

        vendors = get_unique_vendors(helper_email)
        if receipt['vendor'] and receipt['vendor'] not in vendors:
            vendors = [receipt['vendor']] + vendors
        vendor_index = vendors.index(receipt['vendor']) if receipt['vendor'] in vendors else 0

        new_vendor = st.selectbox(
            "取引先",
            options=vendors if vendors else [""],
            index=vendor_index if vendors else 0
        )

        categories = get_categories()
        category_index = categories.index(receipt['category']) if receipt['category'] in categories else 0

        new_category = st.selectbox(
            "費目",
            options=categories,
            index=category_index
        )

        new_description = st.text_area(
            "備考",
            value=receipt['description'] or "",
            height=100
        )

        change_reason = st.text_input(
            "変更理由 (必須)",
            placeholder="変更内容と理由を記述してください（監査ログに記録されます）",
            help="電子帳簿保存法対応のため、変更理由の記録が必須です"
        )

        # ボタンは縦積み
        submit_update = st.form_submit_button("変更を保存", use_container_width=True, type="primary")
        submit_delete = st.form_submit_button("論理削除", use_container_width=True, type="secondary")

        if submit_update:
            if not change_reason:
                st.error("変更理由を入力してください")
            else:
                try:
                    ok = update_receipt(
                        receipt_id=receipt_id,
                        helper_email=helper_email,
                        transaction_date=new_date.isoformat(),
                        amount=float(new_amount),
                        vendor=new_vendor,
                        category=new_category,
                        description=new_description,
                        reason=change_reason,
                        updated_by=helper_email,
                    )

                    if ok:
                        st.success("✅ レシート情報を更新しました")
                        st.rerun()
                    else:
                        st.error("更新できませんでした (所有者が違うかレシートが存在しません)")

                except Exception as e:
                    st.error(f"エラー: {str(e)}")

        if submit_delete:
            if not change_reason:
                st.error("削除理由を入力してください")
            else:
                try:
                    ok = soft_delete_receipt(
                        receipt_id=receipt_id,
                        helper_email=helper_email,
                        reason=change_reason,
                        deleted_by=helper_email,
                    )

                    if ok:
                        st.success("✅ レシートを論理削除しました")
                        st.rerun()
                    else:
                        st.error("削除できませんでした (所有者が違うかレシートが存在しません)")

                except Exception as e:
                    st.error(f"エラー: {str(e)}")

    # ===== 監査ログ表示 =====
    st.divider()
    st.subheader("変更履歴")

    audit_logs = get_audit_log(helper_email=helper_email, receipt_id=receipt_id)

    if audit_logs:
        audit_df = pd.DataFrame(audit_logs)
        audit_df_display = audit_df[[
            'changed_at', 'action', 'changed_column', 'old_value', 'new_value', 'changed_by', 'reason'
        ]].copy()

        audit_df_display.columns = ['日時', '操作', '項目', '変更前', '変更後', '変更者', '理由']

        st.dataframe(
            audit_df_display,
            use_container_width=True,
            hide_index=True
        )
    else:
        st.info("変更履歴がありません")


# ===== ページ5 (管理者専用): 税理士提出 =====
def page_tax_report(user: dict):
    """管理者専用ページ: 全ヘルパーのレシートをまとめて税理士提出用Excelに出力"""
    st.title("📊 税理士提出 (管理者)")

    helper_email = user["helper_email"]

    # 権限チェック (多重防御)
    if not is_admin(helper_email):
        st.error("❌ このページは管理者のみ利用できます。")
        return

    st.write("""
    全ヘルパーの経費レシートをまとめて取得し、税理士提出用のExcelファイルを作成します。
    出力には「登録者」列が含まれ、ヘルパー別の費用を確認できます。
    """)

    # 検索パネル (スマホ対応: 1カラム縦積み)
    with st.expander("🔎 出力条件", expanded=True):
        # デフォルトは過去1年
        default_from = datetime.now().date() - timedelta(days=365)
        default_to = datetime.now().date()

        date_from = st.date_input(
            "取引年月日：開始",
            value=default_from,
            format="YYYY-MM-DD",
            key="tax_date_from",
        )

        date_to = st.date_input(
            "取引年月日：終了",
            value=default_to,
            format="YYYY-MM-DD",
            key="tax_date_to",
        )

        category_filter = st.selectbox(
            "費目で絞り込み（空白=全費目）",
            options=[""] + get_categories(),
            index=0,
            key="tax_category",
        )

    if st.button("全ヘルパーの経費を集計", use_container_width=True, type="primary"):
        results = search_all_receipts(
            date_from=date_from.isoformat() if date_from else None,
            date_to=date_to.isoformat() if date_to else None,
            category=category_filter if category_filter else None,
        )

        if not results:
            st.info("該当するレシートがありません")
            return

        # 統計情報
        stats = get_all_receipt_stats(
            date_from=date_from.isoformat(),
            date_to=date_to.isoformat(),
        )

        c1, c2, c3 = st.columns(3)
        c1.metric("件数", f"{stats['count']}件")
        c2.metric("合計", format_currency(stats['total']))
        c3.metric(
            "平均",
            format_currency(stats['total'] / stats['count'] if stats['count'] > 0 else 0),
        )

        # 結果テーブル (登録者列を含む)
        st.subheader("集計結果（全ヘルパー）")

        df = pd.DataFrame(results)
        # helper_name が無ければ helper_email で埋める
        if 'helper_name' in df.columns:
            df['登録者'] = df['helper_name'].fillna('').replace('', None)
            if 'helper_email' in df.columns:
                df['登録者'] = df['登録者'].fillna(df['helper_email'])
        elif 'helper_email' in df.columns:
            df['登録者'] = df['helper_email']
        else:
            df['登録者'] = ''

        df_display = df[[
            'id', 'transaction_date', 'category', 'amount', 'vendor', '登録者'
        ]].copy()
        df_display.columns = ['ID', '取引日', '費目', '金額', '取引先', '登録者']
        df_display['金額'] = df_display['金額'].apply(lambda x: format_currency(x))

        st.dataframe(
            df_display,
            use_container_width=True,
            hide_index=True,
        )

        # ヘルパー別サマリ (参考表示)
        st.subheader("ヘルパー別サマリ")
        summary_df = df.copy()
        summary_df['登録者'] = summary_df.get('登録者', '')
        by_helper = summary_df.groupby('登録者').agg(
            件数=('id', 'count'),
            合計=('amount', 'sum'),
        ).reset_index().sort_values('合計', ascending=False)
        by_helper['合計'] = by_helper['合計'].apply(lambda x: format_currency(x))
        st.dataframe(by_helper, use_container_width=True, hide_index=True)

        # Excel出力
        st.subheader("📥 税理士提出用Excel")
        excel_data = create_excel_download(
            results,
            date_from=date_from.isoformat() if date_from else None,
            date_to=date_to.isoformat() if date_to else None,
            include_helper=True,
        )
        filename = f"経費明細_全員_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
        st.download_button(
            label="Excelファイルをダウンロード",
            data=excel_data,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


# ===== ページ4: 監査ログ =====
def page_audit_log(user: dict):
    st.title("📊 監査ログ")

    helper_email = user["helper_email"]

    st.write("""
    電子帳簿保存法対応のための変更履歴です。
    あなたが登録したレシートの登録・更新・削除操作が記録されています。
    """)

    # フィルタオプション (スマホ対応: 1カラム縦積み)
    with st.expander("🔎 フィルタ", expanded=False):
        filter_receipt_id = st.number_input(
            "レシートID（空白で全件表示）",
            value=0,
            min_value=0,
            step=1,
            format="%d"
        )

        filter_action = st.selectbox(
            "操作タイプ",
            options=["", "INSERT", "UPDATE", "DELETE"]
        )

    # ログ取得 (自分のレシートのみ)
    all_logs = get_audit_log(helper_email=helper_email)

    # フィルタ適用
    if filter_receipt_id > 0:
        all_logs = [log for log in all_logs if log['receipt_id'] == filter_receipt_id]

    if filter_action:
        all_logs = [log for log in all_logs if log['action'] == filter_action]

    if all_logs:
        st.metric("ログエントリー数", len(all_logs))

        # DataFrameに変換
        log_df = pd.DataFrame(all_logs)
        log_df_display = log_df[[
            'log_id', 'changed_at', 'receipt_id', 'action', 'changed_column',
            'old_value', 'new_value', 'changed_by', 'reason'
        ]].copy()

        log_df_display.columns = ['ログID', '日時', 'レシートID', '操作', '項目', '変更前', '変更後', '変更者', '理由']

        st.dataframe(
            log_df_display,
            use_container_width=True,
            hide_index=True
        )

    else:
        st.info("ログエントリーがありません")


# ===== メインアプリケーション =====
def main():
    # 認証 (失敗時は以降の処理をスキップ)
    if not login_page():
        st.stop()

    user = _get_current_user()
    helper_email = user["helper_email"]
    helper_name = user.get("helper_name", "")

    # サイドバーナビゲーション
    st.sidebar.title("🧾 経費精算・レシート管理")

    # 現在のユーザー情報
    st.sidebar.markdown(f"""
    **ログイン中:**
    {helper_name or helper_email}
    """)
    st.sidebar.caption(helper_email)
    st.sidebar.divider()

    # 管理者のみ「税理士提出」ページを表示
    admin_user = is_admin(helper_email)
    page_options = [
        "📝 レシート登録",
        "⌨️ 手入力",
        "🔍 検索・一覧",
        "📋 詳細・編集",
        "📊 監査ログ",
    ]
    if admin_user:
        page_options.append("📊 税理士提出 (管理者)")

    page = st.sidebar.radio(
        "ページを選択",
        options=page_options,
    )

    if admin_user:
        st.sidebar.success("👑 管理者権限")

    st.sidebar.divider()

    # 統計情報をサイドバーに表示 (自分のレシートのみ)
    st.sidebar.subheader("📈 あなたの統計")

    stats = get_receipt_stats(helper_email=helper_email)
    col1, col2 = st.sidebar.columns(2)

    with col1:
        st.metric("総件数", f"{stats['count']}件")

    with col2:
        st.metric("総金額", format_currency(stats['total']))

    st.sidebar.divider()
    st.sidebar.caption("""
    🔐 電子帳簿保存法対応
    - 全操作が監査ログに記録されます
    - 物理削除は禁止（論理削除のみ）
    - 編集時は理由の記録が必須です
    - レシートは登録者本人のみ閲覧・編集可能
    """)

    st.sidebar.divider()
    if st.sidebar.button("🚪 ログアウト"):
        sign_out_helper()
        st.session_state.pop("current_user", None)
        st.rerun()

    # ページ表示
    if page == "📝 レシート登録":
        page_receipt_registration(user)
    elif page == "⌨️ 手入力":
        page_manual_entry(user)
    elif page == "🔍 検索・一覧":
        page_search_and_list(user)
    elif page == "📋 詳細・編集":
        page_detail_and_edit(user)
    elif page == "📊 監査ログ":
        page_audit_log(user)
    elif page == "📊 税理士提出 (管理者)":
        page_tax_report(user)


if __name__ == "__main__":
    main()
